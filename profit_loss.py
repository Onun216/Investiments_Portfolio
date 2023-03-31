import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from sectores import companies_positions as bought
from sectores import sold_positions as sold

file = load_workbook('Investimentos.xlsx', data_only=True)
file_sheets = file.sheetnames


def profit_loss():
    bought_value = bought()
    sold_value = abs(sold())
    if sold_value > bought_value:
        profit = sold_value - bought_value
        percentage = sold_value * 100 / bought_value
        percentage_profit = abs(100 - percentage)
        print(f'Vendido com lucro de €{round(profit, 2)} : +{round(percentage_profit, 2)}%')
    elif sold_value == 0:
        print('Ainda não vendeu a posição na empresa.')
    else:
        loss = bought_value - sold_value
        percentage = sold_value * 100 / bought_value
        percentage_loss = abs(100 - percentage)
        print(f'Vendido com perda de €{round(loss, 2)} : -{round(percentage_loss, 2)}%')


profit_loss()
