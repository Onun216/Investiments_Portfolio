import openpyxl
import plotly.express as px
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from plotly.graph_objs import Scattergeo, Layout
from plotly import offline

file = load_workbook('Empresas_GPS.xlsx', data_only=True)
file_sheets = file.sheetnames

names, details = [], []
lons, lats = [], []


def get_names():
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column

        for row in range(2, fs_count_row + 1):
            for column in range(1, 2):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                names.append(cell_value)

        return names


def get_detail():
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column

        for row in range(2, fs_count_row + 1):
            for column in range(2, 3):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                details.append(cell_value)

        return details


def get_lat():
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column

        for row in range(2, fs_count_row + 1):
            for column in range(3, 4):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                lats.append(cell_value)

        return lats


def get_lon():
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column

        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                lons.append(cell_value)

        return lons


get_names()
get_detail()
get_lat()
get_lon()

my_layout = Layout(title='Global portfolio distribution')

data = [{
    'type': 'scattergeo',
    'lon': lons,
    'lat': lats,
    'text': names,
    'marker': {
        'size': 10
    }
}]

fig = {'data': data, 'layout': my_layout}
offline.plot(fig, filename='portfolio_distribution.html')


