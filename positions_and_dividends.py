import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

file = load_workbook('Investimentos.xlsx', data_only=True)
file_sheets = file.sheetnames

dividend_color = 'FFFFFF00'  # yellow
enter_position_color = 'FF00FF00'  # green
close_position_color = 'FFFF0000'  # red
total_dividends = []
total_position = []


# sector_positions = {
# 'ETF': [],
# 'Semiconductors': [],
# 'Basic Materials': [],
# 'Financial': [],
# 'Services': [],
# 'Consumer Cyclical': [],
# 'Goods': [],
# 'Transportation': [],
# 'Technology': [],
# 'REIT': [],
# 'Utilities': [],
# 'Energy': []
# }


def get_dividends():
    ano = 2022
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        dividends = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if bgColor == dividend_color:
                    dividends.append(cell_value)
                total = sum(dividends)
        total_dividends.append(total)
        print(f'Dividendos de %s: €{total:.2f}' % ano)
        ano += 1
    print(f'Total: €{round(sum(total_dividends), 2)}')
    print('------------------------------------')


def get_posicoes():
    ano = 2022
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        position = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if fs.cell(column, 4).fill.bgColor.index == dividend_color:
                    continue
                elif fs.cell(column, 4).fill.bgColor.index == close_position_color:
                    continue
                elif isinstance(cell_value, float):
                    position.append(cell_value)
                total = round(sum(position), 2)
        total_position.append(total)
        print(f'Posição em %s: €{total_position}' % ano)
        ano += 1

    print(f'Total: €{round(sum(total_position), 2)}')


cash_position = input("Insira o valor actual disponível para investir: ")


def total_positions_and_dividends(cash_position):
    ano = 2022
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        dividends = []
        position = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if not isinstance(cell_value, float):
                    continue
                # if bgColor == close_position_color:
                # continue
                if bgColor == dividend_color:
                    dividends.append(cell_value)
                total_div = sum(dividends)
                if isinstance(cell_value, float) and not (bgColor == dividend_color):
                    position.append(cell_value)
                total_pos = round(sum(position), 2)

        total_dividends.append(total_div)
        print(f'Dividendos de %s: €{total_div:.2f}' % ano)

        # total_position.append(total_pos)
        # print(f'Posição em %s: €{total_position}' % ano)
        # print(position)

        ano += 1
    # print(f'Custo actual das posições: €{round(sum(total_position), 2)}')
    print(f'Total (dividendos): €{round(sum(total_dividends), 2)}')
    print(f'Total (CASH): €{cash_position}')
    print('------------------------------------')


total_positions_and_dividends(cash_position)
