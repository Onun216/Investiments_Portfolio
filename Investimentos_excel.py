import openpyxl

file = openpyxl.load_workbook('Investimentos.xlsx', data_only=True)
fs = file.active
fs_count_row = fs.max_row
fs_count_col = fs.max_column
dividend_color = 'FFFFFF00'  # yellow
enter_position_color = 'FF00FF00'  # green
close_position_color = ''  # red


    




