import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

file = load_workbook('Investimentos.xlsx', data_only=True)
file_sheets = file.sheetnames
dividend_color = 'FFFFFF00'  # yellow
close_position_color = 'FFFF0000'  # red


def portfolio_sectors():
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        sectors_set = set([])
        for row in range(2, fs_count_row + 1):
            for column in range(3, 4):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                if not isinstance(cell_value, str):
                    continue
                sectors_set.add(cell_value)
        print(list(sectors_set))


def portfolio_companies():
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        companies_set = set([])
        for row in range(2, fs_count_row + 1):
            for column in range(2, 3):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                if not isinstance(cell_value, str):
                    continue
                cell_value = cell_value.strip()
                companies_set.add(cell_value)
        print(list(companies_set))


def companies_sectors():
    ano = 2022
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        companies_sct = {}
        for row in range(2, fs_count_row + 1):
            for column in range(3, 4):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                sector = fs.cell(row, 3).value
                company = fs.cell(row, 2).value
                lst_company = []
                if not isinstance(sector, str):
                    continue
                company = company.strip()
                if sector in companies_sct:
                    if not company in companies_sct[sector]:
                        companies_sct[sector].append(company)
                else:
                    companies_sct[sector] = lst_company

        print(f'Posição por sector em %s: {companies_sct}' % ano)
        ano += 1


def companies_positions():
    company_name = input('Empresa: ')
    ano = 2022
    totals = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        totals_bought = {}
        for row in range(2, fs_count_row + 1):
            for column in range(2, 4):
                cell = fs.cell(column=column, row=row)
                company = fs.cell(row, 2).value
                position = fs.cell(row, 4).value
                lst_positions = []
                if not isinstance(company, str):
                    continue
                company = company.strip()
                if fs.cell(row, 4).fill.bgColor.index == dividend_color:
                    continue
                if fs.cell(row, 4).fill.bgColor.index == close_position_color:
                    continue
                if company in totals_bought:
                    if not position in totals_bought[company]:
                        totals_bought[company].append(position)
                else:
                    totals_bought[company] = lst_positions

        if company_name == 'ALL':
            print(f'Posições de %s: {totals_bought}' % ano)
            ano += 1
        else:
            try:
                total = float([round(sum(totals_bought[company_name]), 2)][0])
                print(f'Posição comprada {company_name} (%s): {totals_bought[company_name]} Total = {total}' % ano)
                totals.append(total)
                ano += 1
            except (KeyError, KeyboardInterrupt):
                    print()
    soma = round(sum(totals), 2)
    print(f'Total da posição actual: €{soma}')
    return soma


def sold_positions():
    company_name = input('Empresa: ')
    ano = 2022
    totals = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        totals_sold = {}
        cnt = 1
        for row in range(2, fs_count_row + 1):
            for column in range(2, 4):
                cell = fs.cell(column=column, row=row)
                company = fs.cell(row, 2).value
                position = fs.cell(row, 4).value
                lst_positions = []
                if not isinstance(company, str):
                    continue
                company = company.strip()
                if fs.cell(row, 4).fill.bgColor.index == dividend_color:
                    continue
                if fs.cell(row, 4).fill.bgColor.index == close_position_color:
                    if company in totals_sold:
                        if not position in totals_sold[company]:
                            totals_sold[company].append(position)
                    else:
                        totals_sold[company] = lst_positions

        if company_name == 'ALL':
            print(f'Posições de %s: {totals_sold}' % ano)
            ano += 1
        else:
            try:
                total = float([round(sum(totals_sold[company_name]), 2)][0])
                while total == 0:
                    cnt += 1
                    continue
                else:
                    print(f'Posição vendida {company_name} (%s): {totals_sold[company_name]} Total = {total}' % (ano + cnt))
                    totals.append(total)
                    ano += 1
            except (KeyError, KeyboardInterrupt):
                print()
    soma_sold_positions = round(sum(totals), 2)
    print(f'Posição final de venda: €{soma_sold_positions}')
    return soma_sold_positions

# portfolio_sectors()
# portfolio_companies()
# companies_sectors()

# companies_positions()
# sold_positions()
